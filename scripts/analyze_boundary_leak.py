#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import re
import difflib
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl


@dataclass(frozen=True)
class Example:
    sentence_id: int
    kind: str
    base: float
    best: float
    delta: float
    length_diff: int
    gt_prev: str
    gt_curr: str
    gt_next: str
    pred: str


def _norm_basic(s: str) -> str:
    return (s or "").strip()


def _norm_space_punct(s: str) -> str:
    s = _norm_basic(s)
    s = re.sub(r"\s+", "", s)
    # Keep: ASCII alnum + Hangul + CJK ideographs
    s = re.sub(r"[^0-9A-Za-z\uAC00-\uD7A3\u4E00-\u9FFF\u3400-\u4DBF]", "", s)
    return s


def _ratio(a: str, b: str) -> float:
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def load_gt_sources(xlsx_path: Path) -> list[str]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {name: idx for idx, name in enumerate(header)}
    src_idx = col_map.get("원문")
    if src_idx is None:
        raise RuntimeError(f"'원문' column not found. header={header}")

    sources: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sources.append(((row[src_idx] or "").strip()))
    return sources


def load_translation_exact_sentence_ids(
    detail_csv: Path,
    *,
    exact_col: str | None = None,
    threshold: float = 0.999,
) -> tuple[set[int], set[int]]:
    """Load sentence_id set where translation (target) is exact.

    detail_csv: 문장별 상세결과 CSV (보통 *_번역기준.csv)

    기본 동작:
    - exact_col 미지정이면 아래 후보 중 존재하는 첫 컬럼을 사용
      1) target_text_match_strict
      2) target_text_match
      3) target_text_similarity
    - 값이 threshold 이상이면 '완전일치'로 간주
    """
    df = pd.read_csv(detail_csv)
    if "sentence_id" not in df.columns:
        raise RuntimeError(f"'sentence_id' column not found in {detail_csv}")

    candidates = [
        exact_col,
        "target_text_match_strict",
        "target_text_match",
        "target_text_similarity",
    ]
    chosen: str | None = None
    for c in candidates:
        if c and c in df.columns:
            chosen = c
            break
    if not chosen:
        raise RuntimeError(
            "No suitable translation-exact column found. "
            f"Provide --tgt-exact-col. columns={list(df.columns)}"
        )

    if "matched_pred_id" not in df.columns:
        # 구버전/다른 포맷 대비: pred id 없이도 GT id만은 필터 가능하게
        df["matched_pred_id"] = pd.NA

    gt_ids: set[int] = set()
    pred_ids: set[int] = set()
    for r in df[["sentence_id", "matched_pred_id", chosen]].itertuples(index=False):
        gt_sid_raw, pred_sid_raw, v_raw = r
        if pd.isna(gt_sid_raw):
            continue
        try:
            gt_sid = int(float(gt_sid_raw))
        except (TypeError, ValueError):
            continue

        if pd.isna(v_raw):
            continue
        try:
            v = float(v_raw)
        except (TypeError, ValueError):
            continue

        if v >= threshold:
            gt_ids.add(gt_sid)
            if pd.notna(pred_sid_raw):
                try:
                    pred_ids.add(int(float(pred_sid_raw)))
                except (TypeError, ValueError):
                    pass

    return gt_ids, pred_ids


def analyze_boundary_leak(
    gt_sources: list[str],
    mismatch_csv: Path,
    *,
    allowed_sentence_ids: set[int] | None = None,
    improve_threshold: float = 0.12,
    top_k_examples: int = 10,
    neighbor_window: int = 1,
) -> tuple[dict[str, int], list[Example]]:
    df = pd.read_csv(mismatch_csv)
    # columns: sentence_id,length_diff,similarity,gt_source,pred_source

    kinds = Counter()
    examples: list[Example] = []

    for r in df.itertuples(index=False):
        sid = int(r.sentence_id)
        if allowed_sentence_ids is not None and sid not in allowed_sentence_ids:
            continue
        length_diff = int(r.length_diff)
        pred = str(r.pred_source) if pd.notna(r.pred_source) else ""

        idx = sid - 1
        gt_curr = gt_sources[idx] if 0 <= idx < len(gt_sources) else ""
        gt_prev = gt_sources[idx - 1] if idx - 1 >= 0 else ""
        gt_next = gt_sources[idx + 1] if idx + 1 < len(gt_sources) else ""

        # Use a normalization that is robust to punctuation/space, since boundary errors often involve those.
        pred_n = _norm_space_punct(pred)
        curr_n = _norm_space_punct(gt_curr)
        prev_n = _norm_space_punct(gt_prev)
        next_n = _norm_space_punct(gt_next)

        base = _ratio(pred_n, curr_n)

        cand: dict[str, float] = {"curr": base}

        # Evaluate within a small neighbor window, because many boundary drifts are not strictly ±1
        # once a single boundary is missed.
        w = max(1, int(neighbor_window))
        for d in range(1, w + 1):
            if idx - d >= 0:
                s = _norm_space_punct(gt_sources[idx - d])
                cand[f"prev_{d}"] = _ratio(pred_n, s)
            if idx + d < len(gt_sources):
                s = _norm_space_punct(gt_sources[idx + d])
                cand[f"next_{d}"] = _ratio(pred_n, s)

        # Concatenation candidates for classic boundary leak patterns.
        if gt_prev:
            cand["prev+curr"] = _ratio(pred_n, prev_n + curr_n)
        if gt_next:
            cand["curr+next"] = _ratio(pred_n, curr_n + next_n)

        best_kind = max(cand.items(), key=lambda kv: kv[1])[0]
        best = cand[best_kind]
        delta = best - base

        # classify boundary failure types
        if best_kind == "curr+next" and delta >= improve_threshold:
            kind = "append_next"   # pred contains tail of next segment
        elif best_kind == "prev+curr" and delta >= improve_threshold:
            kind = "prepend_prev"  # pred contains head of previous segment
        elif best_kind == "next_1" and delta >= improve_threshold:
            kind = "shift_to_next"
        elif best_kind == "prev_1" and delta >= improve_threshold:
            kind = "shift_to_prev"
        elif best_kind.startswith("next_") and delta >= improve_threshold:
            # far drift (offset > 1)
            kind = f"shift_to_next_{best_kind.split('_', 1)[1]}"
        elif best_kind.startswith("prev_") and delta >= improve_threshold:
            kind = f"shift_to_prev_{best_kind.split('_', 1)[1]}"
        else:
            kind = "other"

        kinds[kind] += 1

        if kind != "other":
            examples.append(
                Example(
                    sentence_id=sid,
                    kind=kind,
                    base=round(base, 3),
                    best=round(best, 3),
                    delta=round(delta, 3),
                    length_diff=length_diff,
                    gt_prev=gt_prev,
                    gt_curr=gt_curr,
                    gt_next=gt_next,
                    pred=pred,
                )
            )

    # strongest improvements first
    examples.sort(key=lambda e: (e.delta, e.best), reverse=True)
    return dict(kinds), examples[:top_k_examples]


def _clip(s: str, n: int = 90) -> str:
    s = (s or "").replace("\n", " ")
    return s if len(s) <= n else s[:n] + "…"


def _safe_int(x) -> int | None:
    if pd.isna(x):
        return None
    try:
        return int(float(x))
    except (TypeError, ValueError):
        return None


def build_far_shift_summary(
    gt_sources: list[str],
    mismatch_csv: Path,
    *,
    allowed_sentence_ids: set[int] | None,
    neighbor_window: int,
    improve_threshold: float,
    label: str,
) -> pd.DataFrame:
    """Build a compact summary table for later comparison.

    The kind classification logic matches `analyze_boundary_leak`.
    """
    df = pd.read_csv(mismatch_csv)
    w = max(1, int(neighbor_window))

    out_kinds: list[str] = []
    out_len_delta: list[int] = []
    out_gt_len: list[int] = []

    for r in df.itertuples(index=False):
        sid = _safe_int(getattr(r, "sentence_id", None))
        if sid is None:
            continue
        if allowed_sentence_ids is not None and sid not in allowed_sentence_ids:
            continue

        idx = sid - 1
        gt_curr = gt_sources[idx] if 0 <= idx < len(gt_sources) else ""
        gt_prev = gt_sources[idx - 1] if idx - 1 >= 0 else ""
        gt_next = gt_sources[idx + 1] if idx + 1 < len(gt_sources) else ""

        pred = str(getattr(r, "pred_source", "")) if pd.notna(getattr(r, "pred_source", "")) else ""

        pred_n = _norm_space_punct(pred)
        curr_n = _norm_space_punct(gt_curr)
        prev_n = _norm_space_punct(gt_prev)
        next_n = _norm_space_punct(gt_next)

        base = _ratio(pred_n, curr_n)
        cand: dict[str, float] = {"curr": base}

        for d in range(1, w + 1):
            if idx - d >= 0:
                s = _norm_space_punct(gt_sources[idx - d])
                cand[f"prev_{d}"] = _ratio(pred_n, s)
            if idx + d < len(gt_sources):
                s = _norm_space_punct(gt_sources[idx + d])
                cand[f"next_{d}"] = _ratio(pred_n, s)

        if gt_prev:
            cand["prev+curr"] = _ratio(pred_n, prev_n + curr_n)
        if gt_next:
            cand["curr+next"] = _ratio(pred_n, curr_n + next_n)

        best_kind = max(cand.items(), key=lambda kv: kv[1])[0]
        best = cand[best_kind]
        delta = best - base

        if best_kind == "curr+next" and delta >= improve_threshold:
            kind = "append_next"
        elif best_kind == "prev+curr" and delta >= improve_threshold:
            kind = "prepend_prev"
        elif best_kind == "next_1" and delta >= improve_threshold:
            kind = "shift_to_next"
        elif best_kind == "prev_1" and delta >= improve_threshold:
            kind = "shift_to_prev"
        elif best_kind.startswith("next_") and delta >= improve_threshold:
            kind = f"shift_to_next_{best_kind.split('_', 1)[1]}"
        elif best_kind.startswith("prev_") and delta >= improve_threshold:
            kind = f"shift_to_prev_{best_kind.split('_', 1)[1]}"
        else:
            kind = "other"

        ld = _safe_int(getattr(r, "length_diff", None))
        if ld is None:
            ld = 0

        out_kinds.append(kind)
        out_len_delta.append(ld)
        out_gt_len.append(len(curr_n))

    total = len(out_kinds)
    if total == 0:
        return pd.DataFrame(
            columns=[
                "label",
                "kind",
                "count",
                "count_ratio",
                "length_diff_mean",
                "length_diff_median",
                "gt_curr_len_mean",
                "gt_curr_len_median",
            ]
        )

    s = pd.DataFrame({"kind": out_kinds, "length_diff": out_len_delta, "gt_curr_len": out_gt_len})
    g = s.groupby("kind", dropna=False)
    agg = g.agg(
        count=("kind", "size"),
        length_diff_mean=("length_diff", "mean"),
        length_diff_median=("length_diff", "median"),
        gt_curr_len_mean=("gt_curr_len", "mean"),
        gt_curr_len_median=("gt_curr_len", "median"),
    ).reset_index()

    agg.insert(0, "label", label)
    agg["count_ratio"] = agg["count"] / float(total)

    def _sort_key(k: str) -> tuple[int, str]:
        if k in {"append_next", "prepend_prev", "shift_to_next", "shift_to_prev"}:
            return (0, k)
        if k.startswith("shift_to_next_") or k.startswith("shift_to_prev_"):
            return (1, k)
        if k == "other":
            return (3, k)
        return (2, k)

    agg["_sort"] = agg["kind"].map(lambda x: _sort_key(str(x)))
    agg = agg.sort_values(by="_sort").drop(columns=["_sort"])
    return agg


def main() -> int:
    parser = argparse.ArgumentParser(description="경계 누수/shift 유형(append/prepend/shift)을 mismatch CSV로 진단")
    parser.add_argument(
        "--gt-xlsx",
        type=str,
        default=str(
            Path("test_results/bundle_pa_thr0.72_det_afterfix_20251230")
            / "accuracy_eval"
            / "gt_pa_test_100_from_pd_grouped_by_paragraphid.xlsx"
        ),
        help="GT 엑셀 경로(첫 시트에 '원문' 컬럼 필요)",
    )
    parser.add_argument(
        "--mismatch-src",
        type=str,
        default=str(
            Path("test_results/bundle_pa_thr0.72_det_afterfix_20251230")
            / "pa_thr0.72_det_afterfix_accuracy_row_csv"
            / "원문불일치_상세.csv"
        ),
        help="원문기준(source) mismatch CSV 경로(컬럼: sentence_id,length_diff,similarity,gt_source,pred_source)",
    )
    parser.add_argument(
        "--mismatch-tgt",
        type=str,
        default=str(
            Path("test_results/bundle_pa_thr0.72_det_afterfix_20251230")
            / "pa_thr0.72_det_afterfix_accuracy_row_csv"
            / "원문불일치_상세_번역기준.csv"
        ),
        help="번역기준(target) mismatch CSV 경로(컬럼 동일)",
    )
    parser.add_argument("--label", type=str, default="", help="출력 라벨(비교 시 구분용)")
    parser.add_argument("--improve-threshold", type=float, default=0.12)
    parser.add_argument("--top-k", type=int, default=10)
    parser.add_argument(
        "--neighbor-window",
        type=int,
        default=1,
        help="shift 진단 시 고려할 이웃 윈도우(±N). N=1이면 기존처럼 prev/next만 봄.",
    )
    parser.add_argument(
        "--tgt-exact-detail-csv",
        type=str,
        default=None,
        help=(
            "번역문 완전일치 문장만 대상으로 필터링할 때 쓰는 문장별 상세결과 CSV. "
            "(예: *_번역기준.csv; sentence_id + target_text_match_strict/target_text_similarity 컬럼 필요)"
        ),
    )
    parser.add_argument(
        "--tgt-exact-col",
        type=str,
        default=None,
        help="완전일치 판단 컬럼명(미지정 시 target_text_match_strict/target_text_match/target_text_similarity 순으로 자동 탐색)",
    )
    parser.add_argument(
        "--tgt-exact-threshold",
        type=float,
        default=0.999,
        help="완전일치 판단 임계값(기본 0.999)",
    )
    parser.add_argument(
        "--tgt-exact-id-scope",
        type=str,
        default="either",
        choices=["gt", "pred", "either"],
        help=(
            "mismatch CSV의 sentence_id가 GT/PRED 중 무엇을 가리키는지 불명확할 때 범위를 지정. "
            "either(기본): GT sentence_id와 matched_pred_id 둘 다 허용"
        ),
    )
    parser.add_argument(
        "--out-csv",
        type=str,
        default=None,
        help=(
            "요약 통계를 CSV로 저장. 미지정 시 test_results/boundary_leak_summary_<ts>.csv에 저장. "
            "source/target 각각 한 번씩 append 됩니다."
        ),
    )
    args = parser.parse_args()

    gt_xlsx = Path(args.gt_xlsx)
    mismatch_src = Path(args.mismatch_src)
    mismatch_tgt = Path(args.mismatch_tgt)

    gt_sources = load_gt_sources(gt_xlsx)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_csv_path = Path(args.out_csv) if args.out_csv else (Path("test_results") / f"boundary_leak_summary_{ts}.csv")

    allowed_sids: set[int] | None = None
    if args.tgt_exact_detail_csv:
        gt_ids, pred_ids = load_translation_exact_sentence_ids(
            Path(args.tgt_exact_detail_csv),
            exact_col=(str(args.tgt_exact_col) if args.tgt_exact_col else None),
            threshold=float(args.tgt_exact_threshold),
        )
        scope = str(args.tgt_exact_id_scope)
        if scope == "gt":
            allowed_sids = gt_ids
        elif scope == "pred":
            allowed_sids = pred_ids
        else:
            allowed_sids = set(gt_ids)
            allowed_sids.update(pred_ids)
        print(
            "\n[filter] translation-exact only: "
            f"gt_ids={len(gt_ids)} pred_ids={len(pred_ids)} scope={scope} -> allowed={len(allowed_sids)} "
            f"(from {args.tgt_exact_detail_csv})"
        )

    for label, path in [("원문기준(source)", mismatch_src), ("번역기준(target)", mismatch_tgt)]:
        display_label = label
        if args.label:
            display_label = f"{label} | {args.label}"
        kinds, examples = analyze_boundary_leak(
            gt_sources,
            path,
            allowed_sentence_ids=allowed_sids,
            improve_threshold=float(args.improve_threshold),
            top_k_examples=int(args.top_k),
            neighbor_window=int(args.neighbor_window),
        )
        total = sum(kinds.values())
        print("\n=== 경계 누수/시프트 진단:", display_label, "===")
        print("mismatches:", total)
        core_keys = ["append_next", "prepend_prev", "shift_to_next", "shift_to_prev"]
        for k in core_keys:
            if k in kinds:
                print(f"  {k}: {kinds[k]}")

        # show far shifts if any
        far_keys = sorted([k for k in kinds.keys() if k.startswith("shift_to_next_") or k.startswith("shift_to_prev_")])
        for k in far_keys:
            print(f"  {k}: {kinds[k]}")

        if "other" in kinds:
            print(f"  other: {kinds['other']}")

        if examples:
            print("\n대표 예시(정규화 유사도 개선이 큰 순):")
            for i, ex in enumerate(examples, 1):
                print(
                    f"\n[{i}] sid={ex.sentence_id} kind={ex.kind} base={ex.base} -> best={ex.best} (Δ={ex.delta}) lenΔ={ex.length_diff}"
                )
                print("  GT(prev):", _clip(ex.gt_prev))
                print("  GT(curr):", _clip(ex.gt_curr))
                print("  GT(next):", _clip(ex.gt_next))
                print("  PRED    :", _clip(ex.pred))
        else:
            print("(뚜렷한 누수/시프트 개선 케이스가 거의 없음)")

        # Persist summary row(s) for comparisons across runs.
        summary_label = display_label
        summary_df = build_far_shift_summary(
            gt_sources,
            path,
            allowed_sentence_ids=allowed_sids,
            neighbor_window=int(args.neighbor_window),
            improve_threshold=float(args.improve_threshold),
            label=summary_label,
        )
        out_csv_path.parent.mkdir(parents=True, exist_ok=True)
        if out_csv_path.exists():
            summary_df.to_csv(out_csv_path, mode="a", header=False, index=False, encoding="utf-8")
        else:
            summary_df.to_csv(out_csv_path, index=False, encoding="utf-8")
        print(f"[saved] summary_csv: {out_csv_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
